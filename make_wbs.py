"""
WBS + Gantt Chart 생성 스크립트 (컴팩트판, 40행 이내)
취약점 진단 플랫폼 캡스톤 2026-1
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "WBS_Gantt"

# ===== 스타일 =====
title_font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="1F3864")
header_font = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2E75B6")
phase_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
task_font = Font(name="맑은 고딕", size=9)
week_header_font = Font(name="맑은 고딕", size=8, bold=True, color="FFFFFF")

area_fills = {
    "1": PatternFill("solid", fgColor="BF8F00"),
    "2": PatternFill("solid", fgColor="2E75B6"),
    "3": PatternFill("solid", fgColor="C65911"),
    "4": PatternFill("solid", fgColor="548235"),
    "5": PatternFill("solid", fgColor="7030A0"),
    "6": PatternFill("solid", fgColor="C00000"),
}
status_fills = {
    "done":       PatternFill("solid", fgColor="70AD47"),
    "inprogress": PatternFill("solid", fgColor="FFC000"),
    "todo":       PatternFill("solid", fgColor="BFBFBF"),
    "test":       PatternFill("solid", fgColor="9BC2E6"),
}
month_fills = [
    ("3월", PatternFill("solid", fgColor="C65911")),
    ("4월", PatternFill("solid", fgColor="2E75B6")),
    ("5월", PatternFill("solid", fgColor="548235")),
    ("6월", PatternFill("solid", fgColor="7030A0")),
]

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ===== 데이터 (컴팩트) =====
rows = [
    # 영역 1: 관리자 페이지 (4 tasks)
    ("phase", "1", "관리자 페이지 — 가이드라인 관리 (≈40%)", "", "", None, None, None, "1"),
    ("task",  "1.1", "PDF 파서 + DB 적재 (vs_guideline_items 149개)",                 "완료",   100, 4, 6, "done"),
    ("task",  "1.2", "CLI 비교 도구 — 기존 DB vs 신규 PDF (add/mod/del)",              "진행중",  50, 6, 7, "inprogress"),
    ("task",  "1.3", "Gemini CLI 스크립트 자동 증감 루프 (생성/삭제/재작성/재호출)",  "예정",    0, 7, 9, "todo"),
    ("task",  "1.4", "관리자 웹 연동 (PDF 업로드·diff 표시·승인 UI)",                "진행중",  60, 7, 9, "inprogress"),

    # 영역 2: Windows (5 tasks)
    ("phase", "2", "Windows 수집·판정 파이프라인 (≈73%)", "", "", None, None, None, "2"),
    ("task",  "2.1", "Windows 스크립트 82개 (W-01~64 · PC-01~18)",                   "완료",   100, 4, 5, "done"),
    ("task",  "2.2", "UAC 관리자 권한 자동 실행 (Start-Process RunAs)",              "완료",   100, 5, 6, "done"),
    ("task",  "2.3", "수집 파싱 · Null Byte 정제 · DB 저장",                         "완료",   100, 5, 6, "done"),
    ("task",  "2.4", "규칙 1차 + LLM 2차 하이브리드 판정",                            "완료",   100, 6, 7, "done"),
    ("task",  "2.5", "조치 스크립트 UAC 실행 + Gemini 재작성 + 재수집 검증",          "예정",    0, 7, 9, "todo"),

    # 영역 3: Unix (4 tasks)
    ("phase", "3", "Unix/Linux 수집·판정 파이프라인 (≈35%)", "", "", None, None, None, "3"),
    ("task",  "3.1", "Unix 점검 스크립트 U-01~U-67 (67개)",                          "완료",   100, 4, 5, "done"),
    ("task",  "3.2", "sudo 관리자 권한 실행 + 수집 파싱·저장",                       "진행중",  40, 7, 8, "inprogress"),
    ("task",  "3.3", "규칙 1차 + LLM 2차 하이브리드 판정",                            "진행중",  40, 7, 8, "inprogress"),
    ("task",  "3.4", "조치 스크립트 sudo 실행 + Gemini 재작성 루프",                 "예정",    0, 8, 9, "todo"),

    # 영역 4: 웹 (7 tasks)
    ("phase", "4", "웹 — API · UI · 디자인 (≈65%)", "", "", None, None, None, "4"),
    ("task",  "4.1", "DB 스키마 vs_* 테이블 8개 설계",                                "완료",   100, 3, 4, "done"),
    ("task",  "4.2", "Repository CRUD + 인증·권한 (bcrypt·세션)",                    "완료",   100, 4, 5, "done"),
    ("task",  "4.3", "FastAPI 라우트 8개 + Jinja2 템플릿 15개 (데모)",               "완료",   100, 5, 7, "done"),
    ("task",  "4.4", "병렬 수집 Semaphore · sanitize · Quota 재시도",                "완료",   100, 6, 7, "done"),
    ("task",  "4.5", "UI 개편 (진행률 %·신뢰도 제거·판정 UI·scan_id)",              "예정",    0, 7, 8, "todo"),
    ("task",  "4.6", "리포트 PDF/CSV 다운로드 + 이전 진단 비교 화면",                "예정",    0, 7, 9, "todo"),
    ("task",  "4.7", "디자인·스타일 폴리싱 (색상·차트·반응형·a11y)",                "예정",    0, 8, 11, "todo"),

    # 영역 5: 통합 테스트 (4 tasks)
    ("phase", "5", "통합 테스트 (5월 1~2주)", "", "", None, None, None, "5"),
    ("task",  "5.1", "E2E 통합 테스트 — 실제 Windows PC · Linux 서버",               "예정",    0, 8, 9, "test"),
    ("task",  "5.2", "신규 PDF → 스크립트 자동 생성 루프 검증",                      "예정",    0, 9, 9, "test"),
    ("task",  "5.3", "패치 실행 실패 시 Gemini 재작성 루프 검증",                    "예정",    0, 9, 9, "test"),
    ("task",  "5.4", "예외/엣지 케이스 수집 및 수정",                                "예정",    0, 9, 10, "test"),

    # 영역 6: 최종 완성 (4 tasks)
    ("phase", "6", "최종 완성 & 기말발표 (5월 3~4주 · 발표 6월 1일)", "", "", None, None, None, "6"),
    ("task",  "6.1", "LLM 판정 정확도 벤치마크",                                      "예정",    0, 10, 10, "todo"),
    ("task",  "6.2", "최종 보고서 작성",                                             "예정",    0, 10, 11, "todo"),
    ("task",  "6.3", "기말 발표 PPT · 데모 시연 준비 · 리허설",                      "예정",    0, 10, 11, "todo"),
    ("task",  "6.4", "🎯 기말 발표 (6월 1일)",                                        "예정",    0, 12, 12, "todo"),
]

# ===== 레이아웃 =====
weeks = [(m, ["W1","W2","W3","W4"], f) for m, f in month_fills]
total_cols = 6 + 16

# 타이틀 (1행)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
c = ws.cell(row=1, column=1, value="WBS & Gantt — 주통기 취약점 진단 플랫폼 · 중간 4/16 · 기말 6/1")
c.font = title_font; c.fill = title_fill; c.alignment = center
ws.row_dimensions[1].height = 24

# 월 헤더 (2행)
for i in range(1, 7):
    c = ws.cell(row=2, column=i, value=""); c.fill = header_fill; c.border = border
col_cursor = 7
for month, wlist, fill in weeks:
    ws.merge_cells(start_row=2, start_column=col_cursor, end_row=2, end_column=col_cursor+len(wlist)-1)
    c = ws.cell(row=2, column=col_cursor, value=month)
    c.font = header_font; c.fill = fill; c.alignment = center; c.border = border
    col_cursor += len(wlist)

# 컬럼 헤더 (3행)
headers = ["WBS", "작업명", "상태", "진행률", "시작", "종료"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border

col_cursor = 7
for month, wlist, fill in weeks:
    for w in wlist:
        c = ws.cell(row=3, column=col_cursor, value=w)
        c.font = week_header_font; c.fill = fill; c.alignment = center; c.border = border
        col_cursor += 1

ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 18

# 데이터 행
r = 4
def fmt_week(idx):
    m = idx // 4 + 3
    w = idx % 4 + 1
    return f"{m}/W{w}"

for row in rows:
    kind = row[0]
    if kind == "phase":
        _, wbs, title, _s, _p, _gs, _ge, _st, area = row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
        c = ws.cell(row=r, column=1, value=f"  {wbs}. {title}")
        c.font = phase_font; c.fill = area_fills[area]; c.alignment = left; c.border = border
        ws.row_dimensions[r].height = 20
    else:
        _, wbs, title, status_kor, prog, gs, ge, status_en = row
        values = [wbs, title, status_kor, f"{prog}%", fmt_week(gs) if gs is not None else "", fmt_week(ge) if ge is not None else ""]
        for i, v in enumerate(values, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = task_font; c.border = border
            c.alignment = center if i != 2 else left

        status_bg = {"완료":"E2EFDA","진행중":"FFF2CC","예정":"F2F2F2"}.get(status_kor, "FFFFFF")
        ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=status_bg)

        fill = status_fills[status_en]
        if gs is not None and ge is not None:
            for w_idx in range(gs, ge+1):
                col = 7 + w_idx
                c = ws.cell(row=r, column=col, value="")
                c.fill = fill; c.border = border
        for w_idx in range(0, 16):
            col = 7 + w_idx
            ws.cell(row=r, column=col).border = border
        ws.row_dimensions[r].height = 17
    r += 1

# 현재 시점 빨간 세로선 (4월 3주 = index 6)
current_week_col = 7 + 6
for rr in range(2, r):
    cell = ws.cell(row=rr, column=current_week_col)
    cell.border = Border(
        left=Side(border_style="medium", color="C00000"),
        right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom,
    )

# 범례 (1행)
ws.cell(row=r, column=1, value="범례:").font = Font(bold=True, size=9)
legend_items = [("완료","done"),("진행중","inprogress"),("예정","todo"),("테스트","test")]
for i, (label, key) in enumerate(legend_items):
    c = ws.cell(row=r, column=2+i, value=label)
    c.fill = status_fills[key]; c.font = Font(bold=True, size=9); c.alignment = center; c.border = border
c = ws.cell(row=r, column=6, value="◀ 현재: 2026-04-16 (중간발표)")
c.font = Font(bold=True, color="C00000", size=9)
ws.row_dimensions[r].height = 18

# 컬럼 너비
ws.column_dimensions['A'].width = 5.5
ws.column_dimensions['B'].width = 48
ws.column_dimensions['C'].width = 7
ws.column_dimensions['D'].width = 7
ws.column_dimensions['E'].width = 7
ws.column_dimensions['F'].width = 7
for i in range(7, 23):
    ws.column_dimensions[get_column_letter(i)].width = 4

ws.freeze_panes = "G4"

out = "d:/Task/4/취약점진단/WBS_Gantt_취약점진단.xlsx"
wb.save(out)
print(f"Saved: {out} (rows used: {r})")
